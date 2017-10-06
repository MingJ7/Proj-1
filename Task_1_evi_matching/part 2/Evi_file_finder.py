import os
import datetime
import openpyxl
from tkinter import *
import tkinter.filedialog

def get_initial_file_dlg():
    a = Tk() #initialising tk
    a.withdraw() #To forget this window, make it so that it won't appear
    openfilename=tkinter.filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")], initialdir= os.getcwd())
    try:
        fp=open(openfilename,"r") #start the choose file window
        fp.close() #close the file selection window
    except FileNotFoundError as e: #when the user closes the window or clicks cancel
        print(sys.exc_info()[1]) #prints out error message from system
        raise SystemExit #stop running the program
    finally:
        a.destroy() #completely destroy the current tk class, all windows from tk will close
    return openfilename

def change_dir(path):
    if path != os.getcwd():
        if path == "Docs":
            os.chdir("C:\\Users\\{}\\Documents".format(os.getlogin())) #to change dir to documents folder of the current user
            print(os.getcwd())
        else:
            os.chdir(path)
            print("Directory changed to ==>  " + path)

def get_path(name, adate):
    path = "N:\\ITSec\\Daily Requests"
    if isinstance(adate, str): #if date is a string, make it a date
        adate = datemaker(adate)
    path += "\\{}".format(changerFolder[name])
    path += "\\{}".format(adate.year) #going into folder of year, this seems to always be consitant
    for item in os.listdir(path):
        if "." in item: continue #to remove files, so will only find folders
        """to remove year in the folder/file name if present"""
        if short_month[adate.month] in item.lower(): #if the short form name of the month can be found, add that folder into the path straight away
            path += "\\{}".format(item)
            break
        else: #NOTE for this to work, month must be 2 digits like jan would be 01
            iitem = "".join(item.split("{:4d}".format(adate.year), maxsplit=1))
            if iitem != item: #NOTE this line will cuase the program to fail if the month in number sis done without date
                if "{:02d}".format(adate.month) in iitem:
                    path += "\\{}".format(item)
                    break
    for item in os.listdir(path):
        if "." in item: continue #to remove files, so will only find folders
        """to remove year and month in the folder/file name if present"""
        iitem = "".join(item.split("{:4d}".format(adate.year), maxsplit=1)) #remove the year from folder names for higher accuracy in folder finding
        if short_month[adate.month] in iitem.lower(): #also tries to find short form name of month except also checks the day before adding to path
            if "{:02d}".format(adate.day) in iitem:
                path += "\\{}".format(item)
                break
        else:
            iiitem = "".join(item.split("{:2d}".format(adate.month), maxsplit=1)) #removes the month(in numbers) so that the folder select will be the correct
            if "{:02d}".format(adate.day) in iitem:
                path += "\\{}".format(item)
                break
    return path

def get_file(user, folderpath):
    resultlist = []
    for file in os.listdir(folderpath):
        if file[-4:] == ".pdf": #check if file extention is .pdf
            if user in file: #check if it is the file related to the change, if it is, add to list
                resultlist.append(file)
    if len(resultlist) == 1: #return file name if there is only one file
        return str(resultlist[0])
    elif len(resultlist) > 1: #for more than 1 or zero return a code
        return 2
    elif len(resultlist) == 0:
        return 0

def datemaker(datestr): #convert the date string into a date format
    datestr = datestr.split(".")
    adate = datetime.date(int(datestr[2]), int(datestr[1]), int(datestr[0]))
    return adate

def special_case(user, changer):
    if changer == "" or changer == "":
        return True
    elif user == "" and changer == "":
        if action == "Password changed":
            row[0].value = "reset"
            print(" gotten")
        else:
            row[0].value = "ACTION NEEDED"
            row[0].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=openpyxl.styles.colors.BLACK)
            row[0].font = openpyxl.styles.Font(name="Calibri", b=True, i=False, underline="single", color=openpyxl.styles.colors.RED)
        return True
    elif user == "" and changer == "":
        if action == "Password changed":
            row[0].value = "reset"
            print(" gotten")
        else:
            row[0].value = "ACTION NEEDED"
            row[0].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=openpyxl.styles.colors.BLACK)
            row[0].font = openpyxl.styles.Font(name="Calibri", b=True, i=False, underline="single", color=openpyxl.styles.colors.RED)
        return True
    elif user == "" and changer == "":
        if action == "Password changed":
            row[0].value = "reset"
            print(" gotten")
        else:
            row[0].value = "ACTION NEEDED"
            row[0].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=openpyxl.styles.colors.BLACK)
            row[0].font = openpyxl.styles.Font(name="Calibri", b=True, i=False, underline="single", color=openpyxl.styles.colors.RED)
        return True
    elif user == changer:
        row[0].value = "N/A"
        row[0].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="00FF9900")
        #fill with Orange
        return True

class column_headers:
    userh = "User" #NOTE CHANGE THESE SHOULD THE HEADERS CHANGE
    dateh = "Date"
    changeh = "Changed"
    actionh = "Action"
    def __init__(self,row1):
        self.userCol = None #to store the column
        self.dateCol = None
        self.changeByCol = None
        self.actionCol = None
        for cell in row1:
            if cell.value is not None: #Assign the respective value of the column if the text in cell matches the header
                if self.userh in cell.value:
                    self.userCol = cell.column
                    print("{}{} is user".format(self.userCol,cell.row))
                elif self.dateh == cell.value:
                    self.dateCol = cell.column
                    print("{}{} is date".format(self.dateCol,cell.row))
                elif self.changeh in cell.value:
                    self.changeByCol = cell.column
                    print("{}{} is change".format(self.changeByCol,cell.row))
                elif self.actionh in cell.value:
                    self.actionCol = cell.column
                    print("{}{} is action".format(self.actionCol,cell.row))

short_month = {
    1: "jan",
    2: "feb",
    3: "mar",
    4: "apr",
    5: "may",
    6: "jun",
    7: "jul",
    8: "aug",
    9: "sep",
    10: "oct",
    11: "nov",
    12: "dec"
}

#code START HERE
change_dir("Docs")#changes directory to documents
initialFilePath = get_initial_file_dlg() #starts window to select file and store the file path into initailfilepath
initialFilePath = initialFilePath.split("/") #seperate the path into a list of "folders"
filename = initialFilePath.pop(-1) #get the file name to open (last element in the path and therfore the list)
filepath = "/".join(initialFilePath) #join the rest of the path back together to get the folderpath
change_dir(filepath) #change the directory if filepath is changed from documents
print("path= " + filepath + "++name= " + filename)

try:    #Open Excel document
    mainexcel = openpyxl.load_workbook(filename)
except Exception as error: #This should not happen
    print("File not Found")
    raise SystemExit

for sheet in mainexcel.get_sheet_names(): #to cycle through all excel sheets
    cursheet = mainexcel.get_sheet_by_name(sheet) #changing the current active excel sheet
    colh = column_headers(cursheet[1]) #get the columns for specific headers

    p_user = None #Used to look for evi only at the top line of the list of actions
    p_changer = None
    for row in cursheet.rows: #in each row do below
        if row[0].row == 1 or row[0].row == 2: continue #Skip the 1st 2 rows
        user = None #Have these variables
        adate = None
        changer = None
        action = None
        for cell in row: #Check each cell column and assign data to respective varible if it matches
            if cell.column == colh.dateCol:
                adate = cell.value
            elif cell.column == colh.userCol:
                user = cell.value
            elif cell.column == colh.changeByCol:
                changer = cell.value
            elif cell.column == colh.actionCol:
                action = cell.value
        if user == p_user and changer == p_changer: continue #skip trying to check if change is the same user as previous row and changer is also the same
        else:
            p_user = user
            p_changer = changer
        if special_case(user,changer): continue #Check if it is speacial case, if it is, skip the row after doing code in specail case
        path = get_path(changer, adate) #get the path to the folder, This folder will be the innermost folder the code can get via the date
        file = get_file(user, path) #get the file name if it exists in the path (folder)
        if isinstance(file,str): #if a file is found
            path += "\\{}".format(file) #add the file to path so that the excel file can open the file directly
            row[0].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="9993ff")
            row[0].value = "File" #set the cell text to
        elif file == 2:
            row[0].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="c0ffb2")
            row[0].value = "Multi File"
        elif file == 0:
            row[0].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="fbff42")
            row[0].font = openpyxl.styles.Font(name="Calibri", b=True, color=openpyxl.styles.colors.RED)
            row[0].value = "No File"
        row[0].hyperlink = path
mainexcel.save(filename.split(".")[0] + "-MECHchecked" + ".xlsx") #Needed to keep modifications to excel file
print("Program has finished running")
#NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE NOTE
#BELOW BELOW BELOW BELOW BELOW BELOW BELOW BELOW BELOW BELOW BELOW BELOW BELOW BELOW BELOW
#XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX
""" make the main part of the code a function so that i can call it and spcify the file that i want"""
""" the get_path function now excludes all items with a fullstop in the naem, need to change it so that it will only exclude file with  .xxx"""
"""
2 ways to do the hyperlink:
1. use the excel sheet formula  - i.e. cell.value = "=HYPERLINK("path", "text to be in cell")
    When modifying the value afterwards, excel will autoformat the text such that it will be blue and underline
    otherwise, the text will just be the defualt font
2. set the hyperlink to the cell, meaning text in cell can be changed easily (it will be independent of the hyperlink) - i.e. cell.hyperlink = "path"
    The appearance of the cell in cell won't change after modifying it
    However, all formatting will be removed when hyperlink is removed
Second way has been done for neater looking code
"""
"""
black fill, red words -> should be investigated immediately
orange fill, a N\A has been placed there - This is done in special case
yellow fill, no file has been found - fbff42
light green fill, multiple file has been found -c0ffb2
light blue fill, file link has been done -9993ff
blue fill, linked to file, file checked - code still cannot check contents of file
"""