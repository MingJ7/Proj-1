import os
import datetime
import openpyxl
def change_dir(path):
    if path == "Docs":
        username = os.getlogin() #To get username of current login user to navigate to their documents folder
        os.chdir("C:\\Users\\{}\\Documents".format(username)) #to change dir to documents
        print(os.getcwd())
    else:
        os.chdir(path)
        print("Directory changed to ==>  " + path)
change_dir("Docs")#changes directory to documents
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'].value = "bbb"
fill = openpyxl.styles.PatternFill(fill_type="solid",\
                   start_color=openpyxl.styles.colors.BLACK)
font  = openpyxl.styles.Font(name="Calibri", sz=9, b=True, i=False, underline="single",
                    color=openpyxl.styles.colors.RED)

#ws['A1'].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="c0ffb2")
#ws['A1'].font = openpyxl.styles.Font(name="Calibri", b=True, i=False, underline="single", color=openpyxl.styles.colors.RED)
ws['A1'].hyperlink = "C:\\Users\\{}\\Documents".format(os.getlogin()) #All formatting will be removed when hyperlink is removed
wb.save("kls.xlsx") 

mimicol = \
{
    "orange": "00FF9900",
    "red": "00FF0000",
    "lightblue": "0000FFFF",
    "lightorange": "00FFCC99",
    "purple": "00993366",
    "lightgreen": "75ff7c",
    "lightergreen": "c0ffb2"
}
